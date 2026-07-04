"""نوشتن ردیف‌ها در فایل اکسل با فرمت ۱۷ ستونه‌ی «جدول پرسنل»."""
from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from . import config


def write_workbook(rows: list[dict]):
    """یک Workbook openpyxl می‌سازد و برمی‌گرداند."""
    headers = config.headers()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "جدول پرسنل"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in rows:
        ws.append([row.get(h) for h in headers])

    # عرض ستون‌ها
    widths = {"نام": 12, "نام خانوادگی": 16, "کد پرسنلی": 12, "کد ملی": 14,
              "جنسیت": 14, "شهرستان/منطقه": 18, "نوع قرارداد": 12,
              "نام شرکت": 20, "عنوان شغل": 16, "عنوان لوازم": 16,
              "تاریخ تحویل": 14, "شماره حواله": 14}
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths.get(h, 11)
    ws.freeze_panes = "A2"
    return wb


def to_bytes(rows: list[dict]) -> bytes:
    """خروجی اکسل را به‌صورت بایت برمی‌گرداند (برای دانلود در استریم‌لیت)."""
    wb = write_workbook(rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save(rows: list[dict], path: str) -> str:
    wb = write_workbook(rows)
    wb.save(path)
    return path
