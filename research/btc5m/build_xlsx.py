import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

D = json.load(open('/tmp/claude-0/-home-user-Tictactoestevie/8ef0160f-4bcd-5870-bf9e-1bdc78f3e756/scratchpad/data.json'))
WD = ["دوشنبه","سه‌شنبه","چهارشنبه","پنج‌شنبه","جمعه","شنبه","یکشنبه"]
FONT = "Arial"
H = Font(name=FONT, bold=True, color="FFFFFF", size=11)
HF = PatternFill("solid", fgColor="1F4E79")
B = Font(name=FONT, bold=True, size=11)
N = Font(name=FONT, size=10)
BLUE = Font(name=FONT, size=10, color="0000FF")     # ورودیِ ثابت
YEL = PatternFill("solid", fgColor="FFF2CC")
THIN = Border(*[Side(style="thin", color="BFBFBF")]*4)
MONEY = '$#,##0;($#,##0);-'
PCT = '0.0%'

wb = Workbook()

def sheet(name, widths, headers, rtl=True):
    ws = wb.create_sheet(name)
    ws.sheet_view.rightToLeft = rtl
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h); cell.font = H; cell.fill = HF
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    return ws

# ---------- 1. داده خام ----------
ws = sheet("داده خام", [18,8,10,10,12,9,12,8],
           ["زمان ET","ساعت","روز هفته","ماه","تاریخ","جهت","جریان","برد"])
for r, row in enumerate(D["raw"], 2):
    vals = [row[0], row[1], WD[row[2]], row[3], row[4],
            "بالا" if row[5]=="up" else "پایین", row[6], row[7]]
    for cix, v in enumerate(vals, 1):
        cell = ws.cell(r, cix, v); cell.font = N
RAW_LAST = len(D["raw"]) + 1

# ---------- 2. روزانه ----------
cfgs = [(3,20),(4,20),(5,20),(3,50),(4,50),(5,50)]
hdr = ["تاریخ","روز هفته","ماه","سیگنال","برد","درصد برد"] + \
      [f"سود {r} پله / پایه {b}" for r,b in cfgs]
ws = sheet("روزانه", [12,10,9,9,8,10]+[16]*6, hdr)
for r, d in enumerate(D["daily"], 2):
    ws.cell(r,1,d["date"]).font = N
    ws.cell(r,2,WD[d["weekday"]]).font = N
    ws.cell(r,3,d["month"]).font = N
    ws.cell(r,4,d["n"]).font = N
    ws.cell(r,5,d["w"]).font = N
    cell = ws.cell(r,6,f"=IF(D{r}=0,0,E{r}/D{r})"); cell.font=N; cell.number_format=PCT
    for i,(rg,bs) in enumerate(cfgs):
        cell = ws.cell(r,7+i,d[f"p{rg}_{bs}"]); cell.font=N; cell.number_format=MONEY
DAY_LAST = len(D["daily"]) + 1

# ---------- 3. خلاصه ----------
ws = wb.create_sheet("خلاصه", 0); ws.sheet_view.rightToLeft = True
for i,w in enumerate([34,18,18,18,18,18,18],1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws["A1"] = "گزارش یک‌سالهٔ جریان ترکیبی (طلایی + آماری)"
ws["A1"].font = Font(name=FONT, bold=True, size=14)
meta = [("دورهٔ بررسی", f'{D["first"]}  تا  {D["last"]}'),
        ("تعداد روز", round(D["span"])),
        ("کل سیگنال", D["bets"]),
        ("کل برد", D["wins"])]
r = 3
for k,v in meta:
    ws.cell(r,1,k).font = B
    cell = ws.cell(r,2,v); cell.font = N
    r += 1
ws.cell(r,1,"درصد برد").font = B
cell = ws.cell(r,2,"=B6/B5"); cell.font=N; cell.number_format=PCT; r+=1
ws.cell(r,1,"سیگنال در روز (میانگین)").font = B
cell = ws.cell(r,2,f"=B5/B4"); cell.font=N; cell.number_format='0.0'; r+=1
ws.cell(r,1,"سیگنال در هفته").font = B
cell = ws.cell(r,2,f"=B8*7"); cell.font=N; cell.number_format='0.0'; r+=1
ws.cell(r,1,"سیگنال در ماه").font = B
cell = ws.cell(r,2,f"=B8*30"); cell.font=N; cell.number_format='0.0'; r+=2

top = r
ws.cell(r,1,"مقایسهٔ ساختارهای مارتینگل").font = Font(name=FONT,bold=True,size=12)
r += 1
heads = ["ساختار","سود کل","سود ماهانه","سود هفتگی","سود روزانه",
         "بدترین افت","کف مسیر","انفجار","نرخ انفجار","بدترین چرخه","سود ÷ افت"]
for i,h in enumerate(heads,1):
    cell = ws.cell(r,i,h); cell.font=H; cell.fill=HF
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
first_cfg_row = r+1
for i,cfg in enumerate(D["configs"]):
    rr = r+1+i
    col = get_column_letter(7 + cfgs.index((cfg["rungs"], cfg["base"])))
    ws.cell(rr,1,f'{cfg["rungs"]} پله · پایهٔ {cfg["base"]}$').font = B
    c2 = ws.cell(rr,2,f"=SUM(روزانه!{col}2:{col}{DAY_LAST})"); c2.number_format=MONEY
    c3 = ws.cell(rr,3,f"=B{rr}/$B$4*30"); c3.number_format=MONEY
    c4 = ws.cell(rr,4,f"=B{rr}/$B$4*7"); c4.number_format=MONEY
    c5 = ws.cell(rr,5,f"=B{rr}/$B$4"); c5.number_format=MONEY
    c6 = ws.cell(rr,6,cfg["dd"]); c6.number_format=MONEY; c6.font=BLUE
    c7 = ws.cell(rr,7,cfg["low"]); c7.number_format=MONEY; c7.font=BLUE
    c8 = ws.cell(rr,8,cfg["busts"]); c8.font=BLUE
    c9 = ws.cell(rr,9,f"=H{rr}/{cfg['cycles']}"); c9.number_format=PCT
    c10 = ws.cell(rr,10,-cfg["worst_cycle"]); c10.number_format=MONEY; c10.font=BLUE
    c11 = ws.cell(rr,11,f"=IF(F{rr}=0,0,B{rr}/F{rr})"); c11.number_format='0.0'
    for cix in range(2,12):
        cc = ws.cell(rr,cix)
        if cc.font.color is None or cc.font.color.rgb != "000000FF":
            cc.font = N if cix != 1 else B
        cc.border = THIN
note = ws.cell(r+1+len(D["configs"])+1, 1,
    "اعداد آبی خروجی شبیه‌سازی‌اند (مسیر-وابسته: بدترین افت، کف مسیر، تعداد انفجار) — "
    "با فرمول قابل بازتولید نیستند. بقیهٔ سلول‌ها فرمول‌اند و با تغییر شیت «روزانه» به‌روز می‌شوند.")
note.font = Font(name=FONT, size=9, italic=True)
ws.cell(r+1+len(D["configs"])+2, 1,
    "فرض قیمت: ۵۰-۵۰ بدون اسپرد و کارمزد. با اسپرد واقعی این اعداد کوچک‌تر می‌شوند."
    ).font = Font(name=FONT, size=9, italic=True, color="C00000")

# ---------- 4. ساعت ----------
ws = sheet("ساعت", [10,12,12,10,12], ["ساعت ET","سیگنال","در روز","برد","درصد برد"])
for h in range(24):
    r = h+2
    ws.cell(r,1,h).font=N
    ws.cell(r,2,f'=COUNTIF(\'داده خام\'!$B$2:$B${RAW_LAST},A{r})').font=N
    cell=ws.cell(r,3,f"=B{r}/{round(D['span'])}"); cell.font=N; cell.number_format='0.0'
    ws.cell(r,4,f'=SUMIF(\'داده خام\'!$B$2:$B${RAW_LAST},A{r},\'داده خام\'!$H$2:$H${RAW_LAST})').font=N
    cell=ws.cell(r,5,f"=IF(B{r}=0,0,D{r}/B{r})"); cell.font=N; cell.number_format=PCT

# ---------- 5. روز هفته ----------
ws = sheet("روز هفته", [12,12,10,12]+[18]*6,
           ["روز","سیگنال","برد","درصد برد"]+[f"سود {r} پله / پایه {b}" for r,b in cfgs])
for i,d in enumerate(WD):
    r = i+2
    ws.cell(r,1,d).font=B
    ws.cell(r,2,f'=SUMIF(روزانه!$B$2:$B${DAY_LAST},A{r},روزانه!$D$2:$D${DAY_LAST})').font=N
    ws.cell(r,3,f'=SUMIF(روزانه!$B$2:$B${DAY_LAST},A{r},روزانه!$E$2:$E${DAY_LAST})').font=N
    cell=ws.cell(r,4,f"=IF(B{r}=0,0,C{r}/B{r})"); cell.font=N; cell.number_format=PCT
    for j,_ in enumerate(cfgs):
        col = get_column_letter(7+j)
        cell=ws.cell(r,5+j,f'=SUMIF(روزانه!$B$2:$B${DAY_LAST},$A{r},روزانه!${col}$2:${col}${DAY_LAST})')
        cell.font=N; cell.number_format=MONEY

# ---------- 6. ماهانه ----------
months = sorted({d["month"] for d in D["daily"]})
ws = sheet("ماهانه", [12,12,10,12]+[18]*6,
           ["ماه","سیگنال","برد","درصد برد"]+[f"سود {r} پله / پایه {b}" for r,b in cfgs])
for i,m in enumerate(months):
    r = i+2
    ws.cell(r,1,m).font=B
    ws.cell(r,2,f'=SUMIF(روزانه!$C$2:$C${DAY_LAST},A{r},روزانه!$D$2:$D${DAY_LAST})').font=N
    ws.cell(r,3,f'=SUMIF(روزانه!$C$2:$C${DAY_LAST},A{r},روزانه!$E$2:$E${DAY_LAST})').font=N
    cell=ws.cell(r,4,f"=IF(B{r}=0,0,C{r}/B{r})"); cell.font=N; cell.number_format=PCT
    for j,_ in enumerate(cfgs):
        col = get_column_letter(7+j)
        cell=ws.cell(r,5+j,f'=SUMIF(روزانه!$C$2:$C${DAY_LAST},$A{r},روزانه!${col}$2:${col}${DAY_LAST})')
        cell.font=N; cell.number_format=MONEY

# ---------- 7. رشته باخت ----------
ws = sheet("رشته باخت", [14,12,14,26],
           ["طول رشته","تعداد","درصد","توضیح"])
tot = sum(D["runs"].values())
for i,(k,v) in enumerate(sorted(D["runs"].items(), key=lambda x:int(x[0]))):
    r=i+2
    ws.cell(r,1,int(k)).font=N
    ws.cell(r,2,v).font=N
    cell=ws.cell(r,3,f"=B{r}/{tot}"); cell.font=N; cell.number_format=PCT
    if int(k)>=3:
        ws.cell(r,4,"با ۳ پله اینجا منفجر می‌شوی").font=Font(name=FONT,size=9,color="C00000")
r = len(D["runs"])+3
ws.cell(r,1,"انفجارهای پشت‌سرهم (۳ پله)").font=Font(name=FONT,bold=True,size=12); r+=1
for h,x in enumerate(["تعداد انفجار پیاپی","دفعات","ضرر پایه ۲۰","ضرر پایه ۵۰"],1):
    cell=ws.cell(r,h,x); cell.font=H; cell.fill=HF
for i,(k,v) in enumerate(sorted(D["consec"].items(), key=lambda x:int(x[0]))):
    rr=r+1+i
    ws.cell(rr,1,int(k)).font=N; ws.cell(rr,2,v).font=N
    c=ws.cell(rr,3,f"=-140*A{rr}"); c.font=N; c.number_format=MONEY
    c=ws.cell(rr,4,f"=-350*A{rr}"); c.font=N; c.number_format=MONEY

# ---------- 8. دوام سرمایه ----------
ws = sheet("دوام سرمایه", [14,10,10,14,14,14,16],
           ["سرمایه","پایه","پله","نتیجه","تاریخ صفر شدن","شرط تا آن لحظه","بدترین افت"])
for i,s in enumerate(D["surv"]):
    r=i+2
    c=ws.cell(r,1,s[0]); c.font=N; c.number_format=MONEY
    ws.cell(r,2,s[1]).font=N; ws.cell(r,3,s[2]).font=N
    c=ws.cell(r,4,s[3]); c.font=Font(name=FONT,size=10,bold=True,
        color="C00000" if s[3]=="صفر شد" else "008000")
    ws.cell(r,5,s[4]).font=N; ws.cell(r,6,s[5]).font=N
    c=ws.cell(r,7,s[6]); c.font=N; c.number_format=MONEY
    if s[3]=="صفر شد":
        for cix in range(1,8): ws.cell(r,cix).fill = PatternFill("solid", fgColor="FCE4E4")

del wb["Sheet"]
wb.save('/home/user/Tictactoestevie/docs/research/one-year-backtest.xlsx')
print("saved")
