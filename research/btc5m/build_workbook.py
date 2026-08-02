"""
Workbook with values, not formulas.

LibreOffice in this sandbox cannot recalculate even a three-cell file, so a
formula written here would reach the reader with no cached value and could not
be verified before shipping. Every number below is therefore computed in Python
— the same simulation the report came from — and written directly. This is a
frozen historical report; nothing in it is meant to be re-driven by editing an
input cell.
"""
import json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D=json.load(open('/tmp/claude-0/-home-user-Tictactoestevie/8ef0160f-4bcd-5870-bf9e-1bdc78f3e756/scratchpad/data2.json'))
WD=["دوشنبه","سه‌شنبه","چهارشنبه","پنج‌شنبه","جمعه","شنبه","یکشنبه"]
F="Arial"; SPAN=D["span"]
H=Font(name=F,bold=True,color="FFFFFF",size=11); HF=PatternFill("solid",fgColor="1F4E79")
B=Font(name=F,bold=True,size=11); N=Font(name=F,size=10)
GREEN=PatternFill("solid",fgColor="E2EFDA"); RED=PatternFill("solid",fgColor="FCE4E4")
M='$#,##0;($#,##0);-'; P='0.0%'
CFG=[(3,20),(4,20),(5,20),(3,50),(4,50),(5,50)]
VAR=[("old","بدون قانون ۷"),("new","با قانون ۷")]
wb=Workbook()

def sh(name,widths,heads,idx=None):
    ws=wb.create_sheet(name) if idx is None else wb.create_sheet(name,idx)
    ws.sheet_view.rightToLeft=True
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    for i,h in enumerate(heads,1):
        c=ws.cell(1,i,h); c.font=H; c.fill=HF
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.freeze_panes="A2"; return ws

def put(ws,r,cix,v,fmt=None,font=N,fill=None):
    c=ws.cell(r,cix,v); c.font=font
    if fmt: c.number_format=fmt
    if fill: c.fill=fill
    return c

# aggregates
agg_wd={v:defaultdict(lambda:[0,0]+[0.0]*len(CFG)) for v,_ in VAR}
agg_m={v:defaultdict(lambda:[0,0]+[0.0]*len(CFG)) for v,_ in VAR}
for d in D["daily"]:
    for v,_ in VAR:
        for tgt,key in ((agg_wd[v],WD[d["weekday"]]),(agg_m[v],d["month"])):
            tgt[key][0]+=d[f"{v}_n"]; tgt[key][1]+=d[f"{v}_w"]
            for j,(rg,bs) in enumerate(CFG): tgt[key][2+j]+=d[f"{v}_p{rg}_{bs}"]

# ---- خلاصه ----
ws=wb.create_sheet("خلاصه",0); ws.sheet_view.rightToLeft=True
for i,w in enumerate([32,15,15,15,15,14,14,12,12],1):
    ws.column_dimensions[get_column_letter(i)].width=w
ws["A1"]="گزارش یک‌سالهٔ استخر — با و بدون قانون ۷ (بولینگر + RSI)"
ws["A1"].font=Font(name=F,bold=True,size=14)
ws["A2"]=f'دورهٔ {SPAN} روزه · مارتینگل · قیمتِ ۵۰-۵۰ بدون اسپرد و کارمزد'
ws["A2"].font=Font(name=F,size=10,italic=True,color="C00000")
r=4
for v,lab in VAR:
    tt=D["totals"][v]
    put(ws,r,1,lab,font=B)
    put(ws,r,2,"سیگنال"); put(ws,r,3,tt["n"])
    put(ws,r,4,"دقت"); put(ws,r,5,tt["w"]/tt["n"],P)
    put(ws,r,6,"در روز"); put(ws,r,7,round(tt["n"]/SPAN,1),'0.0')
    put(ws,r,8,"در ماه"); put(ws,r,9,round(tt["n"]/SPAN*30))
    r+=1
r+=1
put(ws,r,1,"سود و ریسک به تفکیکِ ساختار",font=Font(name=F,bold=True,size=12)); r+=1
for i,h in enumerate(["ساختار","سود سال","سود ماهانه","سود هفتگی","سود روزانه",
                      "بدترین افت","کف مسیر","انفجار","سود ÷ افت"],1):
    c=ws.cell(r,i,h); c.font=H; c.fill=HF
    c.alignment=Alignment(horizontal="center",wrap_text=True)
for cfg in D["configs"]:
    r+=1; v=cfg["var"]; lab=dict(VAR)[v]; fill=GREEN if v=="new" else None
    put(ws,r,1,f'{cfg["rungs"]} پله · {cfg["base"]}$ · {lab}',font=B,fill=fill)
    put(ws,r,2,cfg["pnl"],M,fill=fill)
    put(ws,r,3,round(cfg["pnl"]/SPAN*30),M,fill=fill)
    put(ws,r,4,round(cfg["pnl"]/SPAN*7),M,fill=fill)
    put(ws,r,5,round(cfg["pnl"]/SPAN),M,fill=fill)
    put(ws,r,6,cfg["dd"],M,fill=fill)
    put(ws,r,7,cfg["low"],M,fill=fill)
    put(ws,r,8,cfg["busts"],fill=fill)
    put(ws,r,9,round(cfg["pnl"]/cfg["dd"],1) if cfg["dd"] else 0,'0.0',fill=fill)
r+=2
put(ws,r,1,f'قانون ۷ در {D["solo7"]:,} پنجره تنها بود و در {D["veto7"]} پنجره خلافِ استخر گفت.',
    font=Font(name=F,size=9,italic=True)); r+=1
put(ws,r,1,"همهٔ اعداد خروجیِ شبیه‌سازیِ کندل‌به‌کندل‌اند و ثابت — این گزارشِ تاریخی است، نه مدلِ زنده.",
    font=Font(name=F,size=9,italic=True))

# ---- روزانه ----
heads=["تاریخ","روز هفته","ماه"]
for v,lab in VAR:
    heads+=[f"سیگنال ({lab})",f"برد ({lab})"]+[f"{rg} پله/{bs}$ ({lab})" for rg,bs in CFG]
ws=sh("روزانه",[12,10,9]+([13,10]+[15]*6)*2,heads)
for r,d in enumerate(D["daily"],2):
    put(ws,r,1,d["date"]); put(ws,r,2,WD[d["weekday"]]); put(ws,r,3,d["month"])
    col=4
    for v,_ in VAR:
        put(ws,r,col,d[f"{v}_n"]); put(ws,r,col+1,d[f"{v}_w"])
        for j,(rg,bs) in enumerate(CFG): put(ws,r,col+2+j,d[f"{v}_p{rg}_{bs}"],M)
        col+=8

# ---- ساعت ----
ws=sh("ساعت",[9]+[12,10,11]*2,["ساعت ET"]+[x for v,l in VAR for x in
      (f"سیگنال ({l})",f"در روز ({l})",f"دقت ({l})")])
for h in range(24):
    r=h+2; put(ws,r,1,h)
    for k,(v,_) in enumerate(VAR):
        n,w=D["hourly"][str(h)][v]
        put(ws,r,2+k*3,n); put(ws,r,3+k*3,round(n/SPAN,1),'0.0')
        put(ws,r,4+k*3,w/n if n else 0,P)

# ---- روز هفته / ماهانه ----
for name,agg,keys in (("روز هفته",agg_wd,WD),
                      ("ماهانه",agg_m,sorted({d["month"] for d in D["daily"]}))):
    heads=[name]+[x for v,l in VAR for x in
           (f"سیگنال ({l})",f"دقت ({l})",f"۳پله/۲۰$ ({l})",f"۳پله/۵۰$ ({l})")]
    ws=sh(name,[12]+[12,10,16,16]*2,heads)
    for i,k in enumerate(keys):
        r=i+2; put(ws,r,1,k,font=B)
        for j,(v,_) in enumerate(VAR):
            a=agg[v][k]; base=2+j*4
            put(ws,r,base,a[0]); put(ws,r,base+1,a[1]/a[0] if a[0] else 0,P)
            put(ws,r,base+2,round(a[2+CFG.index((3,20))]),M)
            put(ws,r,base+3,round(a[2+CFG.index((3,50))]),M)

# ---- رشته باخت ----
ws=sh("رشته باخت",[12]+[12,12]*2,["طول رشته"]+[f"{x} ({l})" for v,l in VAR
      for x in ("تعداد","درصد")])
allk=sorted({int(k) for v,_ in VAR for k in D["runs"][v]})
tot={v:sum(D["runs"][v].values()) for v,_ in VAR}
for i,k in enumerate(allk):
    r=i+2; put(ws,r,1,k)
    for j,(v,_) in enumerate(VAR):
        n=D["runs"][v].get(str(k),0)
        put(ws,r,2+j*2,n); put(ws,r,3+j*2,n/tot[v],P)

# ---- دوام سرمایه ----
ws=sh("دوام سرمایه",[12,16,8,8,14,15,14],
      ["سرمایه","آرایش","پایه","پله","نتیجه","تاریخ صفر شدن","بدترین افت"])
for i,s in enumerate(D["surv"]):
    r=i+2; fill=RED if s[4]=="صفر شد" else None
    put(ws,r,1,s[0],M,fill=fill); put(ws,r,2,dict(VAR)[s[1]],fill=fill)
    put(ws,r,3,s[2],fill=fill); put(ws,r,4,s[3],fill=fill)
    put(ws,r,5,s[4],font=Font(name=F,size=10,bold=True,
        color="C00000" if s[4]=="صفر شد" else "008000"),fill=fill)
    put(ws,r,6,s[5],fill=fill); put(ws,r,7,s[6],M,fill=fill)

del wb["Sheet"]
wb.save('/home/user/Tictactoestevie/docs/research/one-year-backtest.xlsx')
print("saved")
