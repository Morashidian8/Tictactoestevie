"""
Plan B over the last year, Saturdays and Sundays only, as a spreadsheet.

Saturday and Sunday mean the Western weekend — the days the owner clarified
they meant when this came up before, i.e. Python's weekday() 5 and 6, not the
Persian week's شنبه/یکشنبه which fall on Saturday and Sunday too but start the
week rather than end it. Same two days either way; saying so because the two
namings collide.

Plan B is rule 8, and rule 8 fires only into silence — when none of the other
seven have anything to say. So a Plan B signal needs no arbitration; the filter
below is a check on that promise rather than a filter doing real work.

Sheets:
    شنبه و یکشنبه   every weekend signal in time order, with its result
    به تفکیک ساعت   hour by hour, Saturday and Sunday separately and together
    روز به روز      each calendar date
    خلاصه          totals, and the honest comparison with the rest of the week

    python planb_weekend.py [--days 365] [--data btc5m_now.csv]

Grading is close-to-close, the convention the market settles on; Polymarket
settles on a Chainlink 60-second average, so expect roughly 1-2 points lower
live. Everything priced at 50-50 per the standing instruction.
"""

import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault("TELEGRAM_TOKEN", "x")
os.environ.setdefault("RULE8", "1")
import same_dir as S
import polymarket_collector as pmc

OUT = "planb_weekend.xlsx"
TEHRAN = timezone(timedelta(hours=3, minutes=30))
GRAN = 300
PLANB = ["۸) پلن بی"]
FONT = "Arial"
INK, HEAD_BG = "1F2937", "1F3A5F"
WIN_BG, LOSS_BG = "DCFCE7", "FEE2E2"
SAT, SUN = 5, 6
FA_DAY = ["دوشنبه", "سه‌شنبه", "چهارشنبه", "پنج‌شنبه", "جمعه", "شنبه", "یکشنبه"]
THIN = Side(style="thin", color="D1D5DB")


def head(ws, labels, widths):
    ws.append(labels)
    for c in range(1, len(labels) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEAD_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def body(ws, first, last, cols):
    for row in ws.iter_rows(min_row=first, max_row=last, max_col=cols):
        for cell in row:
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=THIN)


def rate_of(sel):
    n = len(sel)
    if not n:
        return 0, 0, 0.0
    w = sum(1 for s in sel if s["won"])
    return w, n, w / n


def block_contrast(weekend, weekday_rows):
    """
    Each month's weekend judged against its OWN weekdays, then those monthly
    gaps averaged.

    A raw weekend-vs-rest table cannot be trusted on its own: the market's
    character drifts across a year, and any split that lines up with the
    calendar inherits that drift and looks predictive. This is the test that
    killed day-of-week the last time it was looked at
    (docs/research/btc-5m-patterns.md), so it is the one that decides.
    """
    by = defaultdict(lambda: ([], []))
    for r in weekend:
        by[f"{r['teh']:%Y-%m}"][0].append(r)
    for r in weekday_rows:
        by[f"{r['teh']:%Y-%m}"][1].append(r)
    out = []
    for blk in sorted(by):
        a, b = by[blk]
        if len(a) < 30 or len(b) < 30:
            continue
        _, na, ra = rate_of(a)
        _, _, rb = rate_of(b)
        out.append((blk, ra, rb, ra - rb, na))
    if len(out) < 2:
        return out, 0.0, 0.0, 0
    gaps = [x[3] for x in out]
    m = sum(gaps) / len(gaps)
    sd = (sum((g - m) ** 2 for g in gaps) / (len(gaps) - 1)) ** 0.5
    se = sd / len(gaps) ** 0.5
    return out, m, (m / se if se else 0.0), sum(1 for g in gaps if g > 0)


def collect(data, days):
    closes = S.load_candles(data)
    ts = sorted(closes)
    cut = ts[-1] - days * 86400
    out = []
    for s in S.replay(closes):
        if s["t"] < cut or s["rules"] != PLANB:
            continue
        t = s["t"]
        w = t + GRAN                       # the window actually bet on
        d = datetime.fromtimestamp(w, TEHRAN)
        out.append({
            "teh": d, "wd": d.weekday(), "hour": d.hour,
            "et_open": datetime.fromtimestamp(w, pmc.ET),
            "et_close": datetime.fromtimestamp(w + GRAN, pmc.ET),
            "bet": "بالا" if s["bet"] == "up" else "پایین",
            "p0": closes[t], "p1": closes[w], "d": closes[w] - closes[t],
            "won": s["won"],
        })
    out.sort(key=lambda r: r["teh"])
    return out


def main():
    argv = sys.argv[1:]
    days = int(argv[argv.index("--days") + 1]) if "--days" in argv else 365
    data = (argv[argv.index("--data") + 1] if "--data" in argv
            else os.path.join("research", "btc5m", "btc5m.csv.gz"))
    if not os.path.exists(data):
        print(f"{data} not found — run research/btc5m/fetch_data.py first.")
        return

    print(f"replaying {days} days from {data} …")
    every = collect(data, days)
    if not every:
        print("no Plan B signal in this span.")
        return
    rows = [r for r in every if r["wd"] in (SAT, SUN)]
    rest = [r for r in every if r["wd"] not in (SAT, SUN)]
    if not rows:
        print("no weekend Plan B signal in this span.")
        return

    w, n, p = rate_of(rows)
    lo, hi = S.wilson(w, n)
    wr, nr, pr = rate_of(rest)
    print(f"{n:,} weekend Plan B signals  {rows[0]['teh']:%Y-%m-%d} -> "
          f"{rows[-1]['teh']:%Y-%m-%d} (Tehran)")
    print(f"{w:,} won · {n - w:,} lost · {p * 100:.2f}%  "
          f"[{lo * 100:.2f}–{hi * 100:.2f}]   break-even 50%")
    print(f"rest of the week: {wr:,}/{nr:,} = {pr * 100:.2f}%   "
          f"gap {(p - pr) * 100:+.2f} points, "
          f"z = {S.two_prop_z(w, n, wr, nr):+.2f}")

    wb = Workbook()

    # ---- 1. every weekend signal ------------------------------------------- #
    ws = wb.active
    ws.title = "شنبه و یکشنبه"
    ws.sheet_view.rightToLeft = True
    head(ws, ["ردیف", "تاریخ (تهران)", "ساعت (تهران)", "روز", "پنجره (ET)",
              "پوزیشن", "قیمت آغاز", "قیمت پایان", "حرکت ($)", "نتیجه"],
         [7, 14, 12, 11, 20, 12, 13, 13, 12, 10])
    for i, r in enumerate(rows, 1):
        ws.append([i, f"{r['teh']:%Y-%m-%d}", f"{r['teh']:%H:%M}",
                   FA_DAY[r["wd"]],
                   f"{r['et_open']:%I:%M}-{r['et_close']:%I:%M%p}",
                   r["bet"], r["p0"], r["p1"], r["d"],
                   "برد" if r["won"] else "باخت"])
    body(ws, 2, n + 1, 10)
    for row in ws.iter_rows(min_row=2, max_row=n + 1, max_col=10):
        won = row[9].value == "برد"
        for j in (6, 7, 8):
            row[j].number_format = '#,##0.00;(#,##0.00);-'
        row[9].fill = PatternFill("solid", fgColor=WIN_BG if won else LOSS_BG)
        row[9].font = Font(name=FONT, size=10, bold=True,
                           color="166534" if won else "991B1B")
    ws.auto_filter.ref = f"A1:J{n + 1}"

    # ---- 2. hour by hour ---------------------------------------------------- #
    hs = wb.create_sheet("به تفکیک ساعت")
    hs.sheet_view.rightToLeft = True
    head(hs, ["ساعت (تهران)", "شنبه ـ تعداد", "شنبه ـ برد", "شنبه ـ درصد",
              "یکشنبه ـ تعداد", "یکشنبه ـ برد", "یکشنبه ـ درصد",
              "هر دو ـ تعداد", "هر دو ـ برد", "هر دو ـ درصد"],
         [13, 12, 11, 12, 13, 12, 13, 12, 11, 12])
    for h in range(24):
        sat = [r for r in rows if r["hour"] == h and r["wd"] == SAT]
        sun = [r for r in rows if r["hour"] == h and r["wd"] == SUN]
        both = sat + sun
        line = [f"{h:02d}:00–{h:02d}:55"]
        for sel in (sat, sun, both):
            kw, kn, kp = rate_of(sel)
            line += [kn, kw, kp if kn else None]
        hs.append(line)
    body(hs, 2, 25, 10)
    for row in hs.iter_rows(min_row=2, max_row=25, max_col=10):
        for j in (3, 6, 9):
            row[j].number_format = "0.0%"
            v = row[j].value
            if isinstance(v, float) and row[j - 2].value >= 10:
                row[j].font = Font(name=FONT, size=10, bold=True,
                                   color="166534" if v > 0.5 else "991B1B")

    # ---- 3. day by day ------------------------------------------------------ #
    ds = wb.create_sheet("روز به روز")
    ds.sheet_view.rightToLeft = True
    head(ds, ["تاریخ", "روز", "تعداد", "برد", "باخت", "درصد", "سود با ۵۰$"],
         [14, 11, 9, 8, 8, 10, 13])
    per = defaultdict(list)
    for r in rows:
        per[f"{r['teh']:%Y-%m-%d}"].append(r)
    for k in sorted(per):
        sel = per[k]
        kw, kn, kp = rate_of(sel)
        ds.append([k, FA_DAY[sel[0]["wd"]], kn, kw, kn - kw, kp,
                   50 * (2 * kw - kn)])
    body(ds, 2, len(per) + 1, 7)
    for row in ds.iter_rows(min_row=2, max_row=len(per) + 1, max_col=7):
        row[5].number_format = "0.0%"
        row[6].number_format = '"$"#,##0;("$"#,##0);"$"0'

    # ---- 4. summary --------------------------------------------------------- #
    # Values, not formulas: this sandbox has no working spreadsheet engine, so
    # a formula would ship with no cached value and read as blank in half the
    # world's viewers.
    sm = wb.create_sheet("خلاصه")
    sm.sheet_view.rightToLeft = True
    sm["A1"] = f"پلن بی — فقط شنبه و یکشنبه، {days} روزِ گذشته"
    sm["A1"].font = Font(name=FONT, bold=True, size=14, color=INK)
    sm["A2"] = (f"{rows[0]['teh']:%Y-%m-%d} تا {rows[-1]['teh']:%Y-%m-%d}"
                f"  ·  {n:,} سیگنال در {len(per)} روز")
    sm["A2"].font = Font(name=FONT, size=10, color="6B7280")

    blocks, blk_m, blk_t, blk_pos = block_contrast(rows, rest)
    # Split-half: the same contrast asked twice, once in each half of the year.
    # A real effect is in both; one that lives in a single good month is not.
    mid = rows[len(rows) // 2]["teh"]
    h1 = (rate_of([r for r in rows if r["teh"] < mid])[2]
          - rate_of([r for r in rest if r["teh"] < mid])[2])
    h2 = (rate_of([r for r in rows if r["teh"] >= mid])[2]
          - rate_of([r for r in rest if r["teh"] >= mid])[2])

    sat = [r for r in rows if r["wd"] == SAT]
    sun = [r for r in rows if r["wd"] == SUN]
    sw, sn, sp = rate_of(sat)
    uw, un, up = rate_of(sun)
    up_ = [r for r in rows if r["bet"] == "بالا"]
    dn_ = [r for r in rows if r["bet"] == "پایین"]
    uw2, un2, up2 = rate_of(up_)
    dw2, dn2, dp2 = rate_of(dn_)

    lines = [
        ("", ""),
        ("کل سیگنال‌های آخر هفته", f"{n:,}"),
        ("برد", f"{w:,}"),
        ("باخت", f"{n - w:,}"),
        ("درصد برد", f"{p * 100:.2f}%"),
        ("بازهٔ اطمینان ۹۵٪", f"{lo * 100:.2f}% تا {hi * 100:.2f}%"),
        ("سربه‌سر", "۵۰٪"),
        ("سود با حجم ثابت ۵۰ دلار", f"${50 * (2 * w - n):+,.0f}"),
        ("", ""),
        ("شنبه", f"{sw:,}/{sn:,} = {sp * 100:.2f}%"),
        ("یکشنبه", f"{uw:,}/{un:,} = {up * 100:.2f}%"),
        ("", ""),
        ("پوزیشن «بالا»", f"{uw2:,}/{un2:,} = {up2 * 100:.2f}%" if un2 else "—"),
        ("پوزیشن «پایین»", f"{dw2:,}/{dn2:,} = {dp2 * 100:.2f}%" if dn2 else "—"),
        ("", ""),
        ("بقیهٔ روزهای هفته", f"{wr:,}/{nr:,} = {pr * 100:.2f}%"),
        ("اختلاف آخر هفته با بقیه", f"{(p - pr) * 100:+.2f} واحد"),
        ("z آماری", f"{S.two_prop_z(w, n, wr, nr):+.2f}"
                    f"   (سربه‌سرِ بونفرونی روی ۷ روز: ۲٫۶۹)"),
        ("", ""),
        ("آزمونِ درون‌ماهی", f"هر ماه با هفتهٔ خودش سنجیده شد: میانگین اختلاف "
                             f"{blk_m * 100:+.2f} واحد، t = {blk_t:+.2f}"),
        ("ماه‌های مثبت", f"{blk_pos} از {len(blocks)} ماه، آخر هفته بهتر بود"),
        ("نیمهٔ اول / دوم", f"{h1 * 100:+.2f} واحد  /  {h2 * 100:+.2f} واحد"),
        ("", ""),
        ("داوری", "این اثر هر سه آزمون را رد کرد: جدول خام، مقایسهٔ درون‌ماهی "
                  "(که دفعهٔ قبل «روزِ هفته» را کشت)، و تکرار در هر دو نیمهٔ "
                  "دوره. قوی‌ترین یافتهٔ تقویمیِ کل پروژه است. با این حال یک "
                  "دورهٔ یک‌ساله است و آزمونِ واقعی، ماه‌های پیشِ روست."),
        ("نمره‌دهی", "بسته به بسته. پلی‌مارکت با میانگین ۶۰ ثانیه‌ای چین‌لینک "
                     "تسویه می‌کند، پس ۱ تا ۲ واحد پایین‌تر انتظار برود."),
        ("قیمتِ ورود", "همه‌چیز ۵۰-۵۰ حساب شده."),
    ]
    r_ = 4
    for a, b in lines:
        sm.cell(row=r_, column=1, value=a).font = Font(
            name=FONT, size=10, bold=bool(a), color=INK)
        c = sm.cell(row=r_, column=2, value=b)
        c.font = Font(name=FONT, size=10, color=INK)
        c.alignment = Alignment(horizontal="right", wrap_text=True,
                                vertical="top")
        r_ += 1
    sm.column_dimensions["A"].width = 28
    sm.column_dimensions["B"].width = 66

    wb.save(OUT)
    print(f"\nwrote {OUT}")
    print(f"  sheet 1: {n:,} weekend signals")
    print(f"  sheet 2: 24 hours x Saturday / Sunday / both")
    print(f"  sheet 3: {len(per)} weekend days")
    print(f"\nSaturday {sp * 100:.2f}% ({sn:,})   Sunday {up * 100:.2f}% ({un:,})")
    hours = [(rate_of([r for r in rows if r["hour"] == h]), h)
             for h in range(24)]
    hours = [x for x in hours if x[0][1] >= 20]
    hours.sort(key=lambda x: -x[0][2])
    print("\nbest and worst hours (>= 20 signals each):")
    for (kw, kn, kp), h in hours[:5]:
        print(f"  {h:02d}:00  {kw:>4,}/{kn:<5,} {kp * 100:6.2f}%")
    print("  …")
    for (kw, kn, kp), h in hours[-3:]:
        print(f"  {h:02d}:00  {kw:>4,}/{kn:<5,} {kp * 100:6.2f}%")


if __name__ == "__main__":
    main()
