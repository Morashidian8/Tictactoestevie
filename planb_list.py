"""
Every Plan B signal of the last two months, in order, with its result.

Plan B is rule 8, and rule 8 speaks only into silence — it fires when all seven
of the other rules have nothing to say. So "a Plan B signal" needs no
arbitration: it is a window where rule 8 was the whole of the signal, and this
list contains those and nothing else.

    python planb_list.py [--days 60] [--data btc5m_now.csv]

Writes planb_signals.xlsx — one sheet in time order, one of totals, one of the
day-by-day record — and prints the summary. Grading is close-to-close, the
convention the market settles on; Polymarket uses a Chainlink 60-second TWAP,
so expect roughly 1-2 points lower live.
"""

import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault("TELEGRAM_TOKEN", "x")
os.environ.setdefault("RULE8", "1")            # the whole point of this file
import same_dir as S
import polymarket_collector as pmc

OUT = "planb_signals.xlsx"
TEHRAN = timezone(timedelta(hours=3, minutes=30))
GRAN = 300
FONT = "Arial"
INK, HEAD_BG = "1F2937", "1F3A5F"
WIN_BG, LOSS_BG = "DCFCE7", "FEE2E2"

HEAD = ["ردیف", "تاریخ (تهران)", "ساعت (تهران)", "روز", "پنجره (ET)",
        "جهت سیگنال", "قیمت آغاز", "قیمت پایان", "حرکت ($)", "نتیجه"]
WIDTH = [7, 14, 12, 11, 20, 12, 12, 12, 12, 10]
FA_DAY = ["دوشنبه", "سه‌شنبه", "چهارشنبه", "پنج‌شنبه", "جمعه", "شنبه", "یکشنبه"]


def collect(data, days):
    closes = S.load_candles(data)
    ts = sorted(closes)
    cut = ts[-1] - days * 86400
    rows = []
    for s in S.replay(closes):
        if s["t"] < cut:
            continue
        # Rule 8 alone, which is what Plan B means. `evaluate` already refuses
        # to let it speak over anything else, so this is a check on that
        # promise rather than a filter doing real work — but a promise that is
        # never checked is how the books quietly merge again.
        if s["rules"] != ["۸) پلن بی"]:
            continue
        t = s["t"]
        w = t + GRAN                 # the window actually bet on
        rows.append({
            "teh": datetime.fromtimestamp(w, TEHRAN),
            "et_open": datetime.fromtimestamp(w, pmc.ET),
            "et_close": datetime.fromtimestamp(w + GRAN, pmc.ET),
            "bet": "بالا" if s["bet"] == "up" else "پایین",
            "p0": closes[t],
            "p1": closes[w],
            "d": closes[w] - closes[t],
            "won": s["won"],
        })
    rows.sort(key=lambda r: r["teh"])
    return rows


def main():
    argv = sys.argv[1:]
    days = int(argv[argv.index("--days") + 1]) if "--days" in argv else 60
    data = (argv[argv.index("--data") + 1] if "--data" in argv
            else os.path.join("research", "btc5m", "btc5m.csv.gz"))
    if not os.path.exists(data):
        print(f"{data} not found — run research/btc5m/fetch_data.py first.")
        return
    print(f"replaying {days} days from {data} …")
    rows = collect(data, days)
    if not rows:
        print("no Plan B signal in this span.")
        return
    won = sum(1 for r in rows if r["won"])
    n = len(rows)
    lo, hi = S.wilson(won, n)
    print(f"{n:,} Plan B signals  {rows[0]['teh']:%Y-%m-%d} -> "
          f"{rows[-1]['teh']:%Y-%m-%d} (Tehran)")
    print(f"{won:,} won  ·  {n - won:,} lost  ·  {won / n * 100:.2f}%  "
          f"[{lo * 100:.2f}–{hi * 100:.2f}]   break-even 50%")

    wb = Workbook()
    ws = wb.active
    ws.title = "پلن بی"
    ws.sheet_view.rightToLeft = True
    ws.append(HEAD)
    for c in range(1, len(HEAD) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEAD_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    for i, r in enumerate(rows, 1):
        ws.append([i, f"{r['teh']:%Y-%m-%d}", f"{r['teh']:%H:%M}",
                   FA_DAY[r["teh"].weekday()],
                   f"{r['et_open']:%I:%M}-{r['et_close']:%I:%M%p}",
                   r["bet"], r["p0"], r["p1"], r["d"],
                   "برد" if r["won"] else "باخت"])
    thin = Side(style="thin", color="D1D5DB")
    for row in ws.iter_rows(min_row=2, max_row=n + 1):
        w_ = row[9].value == "برد"
        for cell in row:
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        for j in (6, 7, 8):
            row[j].number_format = '#,##0.00;(#,##0.00);-'
        row[9].fill = PatternFill("solid", fgColor=WIN_BG if w_ else LOSS_BG)
        row[9].font = Font(name=FONT, size=10, bold=True,
                           color="166534" if w_ else "991B1B")
    for i, w_ in enumerate(WIDTH, 1):
        ws.column_dimensions[get_column_letter(i)].width = w_
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEAD))}{n + 1}"

    # ---- day by day --------------------------------------------------------- #
    per = defaultdict(lambda: [0, 0])
    for r in rows:
        k = f"{r['teh']:%Y-%m-%d}"
        per[k][0] += 1
        per[k][1] += 1 if r["won"] else 0
    d = wb.create_sheet("روز به روز")
    d.sheet_view.rightToLeft = True
    d.append(["تاریخ", "روز", "تعداد", "برد", "باخت", "درصد"])
    for c in range(1, 7):
        cell = d.cell(row=1, column=c)
        cell.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEAD_BG)
        cell.alignment = Alignment(horizontal="center")
    for k in sorted(per):
        tot, w_ = per[k]
        day = datetime.strptime(k, "%Y-%m-%d").weekday()
        d.append([k, FA_DAY[day], tot, w_, tot - w_, w_ / tot])
    for row in d.iter_rows(min_row=2, max_row=len(per) + 1, max_col=6):
        for cell in row:
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=thin)
        row[5].number_format = "0.0%"
    for i, w_ in enumerate((14, 11, 9, 8, 8, 10), 1):
        d.column_dimensions[get_column_letter(i)].width = w_
    d.freeze_panes = "A2"

    # ---- totals, as values ------------------------------------------------- #
    # Values, not formulas: this sandbox's LibreOffice cannot open a
    # spreadsheet at all, so recalc can never run here and openpyxl leaves
    # formulas with no cached value — blank in half the world's readers.
    sm = wb.create_sheet("خلاصه")
    sm.sheet_view.rightToLeft = True
    sm["A1"] = f"پلن بی — همهٔ سیگنال‌های {days} روزِ گذشته"
    sm["A1"].font = Font(name=FONT, bold=True, size=14, color=INK)
    sm["A2"] = (f"{rows[0]['teh']:%Y-%m-%d} تا {rows[-1]['teh']:%Y-%m-%d}"
                f"  ·  {n:,} سیگنال")
    sm["A2"].font = Font(name=FONT, size=10, color="6B7280")
    body = [("", ""), ("کل سیگنال‌ها", f"{n:,}"), ("برد", f"{won:,}"),
            ("باخت", f"{n - won:,}"), ("درصد برد", f"{won / n * 100:.2f}%"),
            ("بازهٔ اطمینان ۹۵٪", f"{lo * 100:.2f}% تا {hi * 100:.2f}%"),
            ("سربه‌سر", "۵۰٪"),
            ("سود با ۵۰ دلار ثابت", f"${50 * (2 * won - n):+,.0f}"), ("", "")]
    up = [r for r in rows if r["bet"] == "بالا"]
    dn = [r for r in rows if r["bet"] == "پایین"]
    for lbl, sel in (("سیگنال‌های «بالا»", up), ("سیگنال‌های «پایین»", dn)):
        if sel:
            k = sum(1 for r in sel if r["won"])
            body.append((lbl, f"{k}/{len(sel)} = {k / len(sel) * 100:.1f}%"))
    body += [("", ""),
             ("نمره‌دهی", "بسته به بسته؛ پلی‌مارکت با میانگینِ ۶۰ ثانیهٔ "
                          "چین‌لینک تسویه می‌کند، پس ۱ تا ۲ واحد پایین‌تر "
                          "انتظار برود."),
             ("قیمتِ ورود", "همه‌چیز ۵۰-۵۰ حساب شده.")]
    r_ = 4
    for a, b in body:
        sm.cell(row=r_, column=1, value=a).font = Font(name=FONT, size=10,
                                                       bold=bool(a), color=INK)
        c = sm.cell(row=r_, column=2, value=b)
        c.font = Font(name=FONT, size=10, color=INK)
        c.alignment = Alignment(horizontal="right", wrap_text=True)
        r_ += 1
    sm.column_dimensions["A"].width = 26
    sm.column_dimensions["B"].width = 62

    wb.save(OUT)
    print(f"\nwrote {OUT}: {n:,} rows over {len(per)} days")
    print(f"\nfirst 5 and last 5:")
    for r in rows[:5] + rows[-5:]:
        print(f"  {r['teh']:%Y-%m-%d %H:%M}  {FA_DAY[r['teh'].weekday()]:<9}"
              f"{r['bet']:<7}{r['d']:>+9.2f}  "
              f"{'برد' if r['won'] else 'باخت'}")


if __name__ == "__main__":
    main()
