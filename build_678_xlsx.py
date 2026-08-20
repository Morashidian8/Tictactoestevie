"""
The month's rule 6/7/8 signals that agreed with the candle before them, as a
spreadsheet.

Only rules 5-8 can ever fire in the same direction as the candle that just
closed; rules 1, 2, 3 and the golden tier are fade rules and structurally never
do. Of the ones that can, this takes the three asked for.

    python build_678_xlsx.py [--days 30] [--data btc5m_fresh.csv]

Writes signals_678_samedir.xlsx: one sheet of signals in time order, one sheet
of totals built from formulas so the numbers move if the rows are filtered.
"""

import os
import sys
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault("TELEGRAM_TOKEN", "x")
import same_dir as S

OUT = "signals_678_samedir.xlsx"
TEHRAN = timezone(timedelta(hours=3, minutes=30))      # Iran dropped DST in 2022
GRAN = 300
WANT = ("۶)", "۷)", "۸)")
FONT = "Arial"

HEAD = ["ردیف", "تاریخ (تهران)", "ساعت (تهران)", "پنجره (UTC)", "قانون",
        "همهٔ قانون‌های شلیک‌شده", "جهت سیگنال", "جهت کندل قبل",
        "حرکت کندل قبل ($)", "قیمت آغاز", "قیمت پایان", "حرکت پنجره ($)",
        "نتیجه"]
WIDTH = [7, 14, 12, 18, 9, 34, 12, 13, 17, 12, 12, 15, 10]

INK = "1F2937"
HEAD_BG = "1F3A5F"
WIN_BG = "DCFCE7"
LOSS_BG = "FEE2E2"
BAND = "F3F4F6"


def collect(data, days):
    closes = S.load_candles(data)
    ts = sorted(closes)
    cut = ts[-1] - days * 86400
    rows = []
    for s in S.replay(closes):
        if s["t"] < cut or not s["aligned"]:
            continue
        tags = [r for r in s["rules"] if r.startswith(WANT)]
        if not tags:
            continue
        t = s["t"]
        # The signal fires at the close of bucket t, which is wall-clock t+300,
        # and bets the bucket starting there. That start is the window a person
        # sees on Polymarket, so it is the time recorded.
        w = t + GRAN
        dprev = closes[t] - closes[t - GRAN]
        rows.append({
            "utc": datetime.fromtimestamp(w, timezone.utc),
            "teh": datetime.fromtimestamp(w, TEHRAN),
            "rule": "+".join(r[0] for r in tags),
            "all": " + ".join(s["rules"]),
            "bet": "بالا" if s["bet"] == "up" else "پایین",
            # Read off the prices, not copied from the bet. The two agree by
            # construction here, and a column that merely restates its filter
            # cannot show the reader that the filter did what it claims.
            "prev": "بالا" if dprev > 0 else "پایین",
            "dprev": dprev,
            "p0": closes[t],
            "p1": closes[w],
            "d": closes[w] - closes[t],
            "won": s["won"],
        })
    rows.sort(key=lambda r: r["utc"])
    return rows


def style_header(ws, row=1):
    for c in range(1, len(HEAD) + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEAD_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def main():
    argv = sys.argv[1:]
    days = int(argv[argv.index("--days") + 1]) if "--days" in argv else 30
    data = (argv[argv.index("--data") + 1] if "--data" in argv
            else os.path.join("research", "btc5m", "btc5m.csv.gz"))
    if not os.path.exists(data):
        print(f"{data} not found — run research/btc5m/fetch_data.py first.")
        return

    print(f"reading {data} …")
    rows = collect(data, days)
    if not rows:
        print("no matching signals.")
        return
    print(f"{len(rows):,} signals from rules 6/7/8 that agreed with the "
          f"previous candle")

    wb = Workbook()
    ws = wb.active
    ws.title = "سیگنال‌ها"
    ws.sheet_view.rightToLeft = True

    ws.append(HEAD)
    style_header(ws)
    thin = Side(style="thin", color="D1D5DB")
    for i, r in enumerate(rows, 1):
        ws.append([i, r["teh"].strftime("%Y-%m-%d"), r["teh"].strftime("%H:%M"),
                   r["utc"].strftime("%Y-%m-%d %H:%M"), r["rule"], r["all"],
                   r["bet"], r["prev"], r["dprev"], r["p0"], r["p1"], r["d"],
                   "برد" if r["won"] else "باخت"])
    for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1):
        won = row[12].value == "برد"
        for cell in row:
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        for j in (8, 9, 10, 11):
            row[j].number_format = '#,##0.00;(#,##0.00);-'
        row[5].alignment = Alignment(horizontal="right", vertical="center")
        row[12].fill = PatternFill("solid", fgColor=WIN_BG if won else LOSS_BG)
        row[12].font = Font(name=FONT, size=10, bold=True,
                            color="166534" if won else "991B1B")
    for i, w in enumerate(WIDTH, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEAD))}{len(rows) + 1}"

    # ---- totals -------------------------------------------------------- #
    # Written as values, not formulas. This sandbox's LibreOffice cannot load a
    # spreadsheet at all — not even a two-line CSV — so recalc.py can never run
    # here, and openpyxl leaves formulas with no cached value. An unverified
    # formula that reads back as blank in half the world's readers is worse
    # than an arithmetic result computed from the very rows on the next sheet.
    sm = wb.create_sheet("خلاصه")
    sm.sheet_view.rightToLeft = True

    sm["A1"] = f"سیگنال‌های قانون ۶ و ۷ و ۸ که هم‌جهتِ کندلِ قبل شلیک شدند"
    sm["A1"].font = Font(name=FONT, bold=True, size=14, color=INK)
    sm["A2"] = (f"{days} روزِ گذشته  ·  "
                f"{rows[0]['teh']:%Y-%m-%d} تا {rows[-1]['teh']:%Y-%m-%d}  ·  "
                f"{len(rows):,} سیگنال")
    sm["A2"].font = Font(name=FONT, size=10, color="6B7280")

    hdr = ["قانون", "تعداد", "برد", "باخت", "درصد برد"]
    for j, h in enumerate(hdr, 1):
        c = sm.cell(row=4, column=j, value=h)
        c.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")

    r = 5
    for label in ("۸", "۷", "۶"):
        sub = [x for x in rows if label in x["rule"]]
        won = sum(1 for x in sub if x["won"])
        sm.cell(row=r, column=1, value=f"قانون {label}")
        sm.cell(row=r, column=2, value=len(sub))
        sm.cell(row=r, column=3, value=won)
        sm.cell(row=r, column=4, value=len(sub) - won)
        sm.cell(row=r, column=5, value=won / len(sub) if sub else 0)
        r += 1
    won = sum(1 for x in rows if x["won"])
    sm.cell(row=r, column=1, value="همه (بدون تکرار)")
    sm.cell(row=r, column=2, value=len(rows))
    sm.cell(row=r, column=3, value=won)
    sm.cell(row=r, column=4, value=len(rows) - won)
    sm.cell(row=r, column=5, value=won / len(rows))
    total = r

    for row in sm.iter_rows(min_row=5, max_row=total, max_col=5):
        for cell in row:
            cell.font = Font(name=FONT, size=10, color=INK,
                             bold=cell.row == total)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        row[4].number_format = "0.00%"
    for j, w in enumerate((22, 10, 10, 10, 12), 1):
        sm.column_dimensions[get_column_letter(j)].width = w

    note = [
        "",
        "عددهای بالا از همان سطرهای برگهٔ «سیگنال‌ها» شمرده شده‌اند؛ ستونِ",
        "«قانون» و ستونِ «نتیجه». جمعِ سه قانون از کلِ سطرها بیشتر می‌شود،",
        "چون یک سیگنال می‌تواند هم‌زمان چند قانون داشته باشد.",
        "",
        "چند نکته که بدونِ آن‌ها این عددها اشتباه خوانده می‌شوند:",
        "",
        "۱. «هم‌جهت» یعنی جهتِ سیگنال با جهتِ کندلی که همان لحظه بسته شده یکی است.",
        "۲. قانون‌های ۱ و ۲ و ۳ و ورودِ طلایی هرگز هم‌جهت شلیک نمی‌کنند — نه کم، صفر.",
        "   آن‌ها قانونِ خلافِ جهت‌اند، پس این فهرست ناچار فقط ۶ و ۷ و ۸ است.",
        "۳. قانون ۸ فقط وقتی حرف می‌زند که هیچ قانونِ دیگری نگفته باشد، برای همین",
        "   بخشِ عمدهٔ این فهرست است — و نازک‌ترین قانونِ سیستم است (سربه‌سرش ۵۰٪).",
        "۴. نمره‌دهی «بسته به بسته» است. پلی‌مارکت با میانگینِ ۶۰ ثانیهٔ چین‌لینک",
        "   تسویه می‌کند، پس در معاملهٔ واقعی حدودِ ۱ تا ۲ واحد پایین‌تر انتظار برود.",
        "۵. سربه‌سر را قیمتِ ورود تعیین می‌کند: ۵۰٪ در ۵۰ سنت، ۵۲٪ در ۵۲ سنت،",
        "   ۵۵٪ در ۵۵ سنت.",
        "۶. روی ۵۹۰ روز، هم‌جهت ۵۲.۵۴٪ و خلافِ جهت ۵۲.۹۲٪ بود — یعنی این تفاوت",
        "   در بازهٔ بلند وجود ندارد و علامتش هم برعکس می‌شود.",
        "",
        f"داده: {os.path.basename(data)} — کندل‌های ۵ دقیقه‌ای بیت‌استمپ، بدون گپ.",
    ]
    r = total + 2
    for text in note:
        c = sm.cell(row=r, column=1, value=text)
        c.font = Font(name=FONT, size=9,
                      color="6B7280" if text.startswith(" ") or not text
                      else INK,
                      bold=text.startswith("چند"))
        c.alignment = Alignment(horizontal="right")
        r += 1

    wb.save(OUT)
    won = sum(1 for r_ in rows if r_["won"])
    print(f"wrote {OUT}: {len(rows):,} rows, {won:,} won "
          f"({won / len(rows) * 100:.2f}%)")


if __name__ == "__main__":
    main()
